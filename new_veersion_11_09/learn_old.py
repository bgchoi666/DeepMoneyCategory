#!/usr/bin/env python
# coding: utf-8

from sklearn import linear_model
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
import numpy as np
import pandas as pd
from sklearn.metrics import confusion_matrix
from numba import jit
from openpyxl import load_workbook

import util, models

import tensorflow as tf
from tensorflow import keras
import math
import os
import operator

class EarlyStopping():
    def __init__(self, patience=0, verbose=0):
        self._step = 0
        self._loss = float('inf')
        self.patience  = patience
        self.verbose = verbose

    def validate(self, loss):
        if self._loss < loss:
            self._step += 1
            print('step : ',self._step)
            if self._step > self.patience:
                if self.verbose:
                    print('Training process is stopped early....')
                return True
        else:
            self._step = 0
            self._loss = loss

        return False

class FeatureSelection(object):
    @staticmethod
    def lasso_fit(X,y, alpha = 1.0, model_return = False):
        model = linear_model.Lasso(alpha=alpha) #모델 생성
        model.fit(X,y)   #모델 학습
        print("#success Lasso model training")
        feature_score = pd.DataFrame({'feature' : X.columns, 'score' : abs(model.coef_)}) #변수 중요도(weight 절대값) 추출
        feature_sorted_score = feature_score.sort_values(by=['score'],ascending = False)  #score 정렬
        if model_return ==True:
            return model,feature_sorted_score
        return feature_sorted_score

    @staticmethod
    def ridge_fit(X,y, alpha = 1.0,model_return = False):
        model = linear_model.Ridge(alpha=alpha)
        model.fit(X,y)
        print("#success Ridge model training")
        feature_score = pd.DataFrame({'feature' : X.columns, 'score' : abs(model.coef_)})
        feature_sorted_score = feature_score.sort_values(by=['score'],ascending = False) 
        if model_return ==True:
            return model,feature_sorted_score
        return feature_sorted_score

    @staticmethod
    def randomforest_classification_fit(X,y,n_estimators = 500, n_jobs = -1,min_samples_leaf = 50, model_return = False):
        model = RandomForestClassifier(n_estimators = n_estimators, n_jobs = n_jobs, min_samples_leaf = min_samples_leaf)
        model.fit(X, y)
        print("#success RF model training")
        feature_score = pd.DataFrame({'feature' : X.columns, 'score' : model.feature_importances_})
        feature_sorted_score = feature_score.sort_values(by=['score'],ascending = False) 
        if model_return ==True:
            return model,feature_sorted_score
        return feature_sorted_score
            
# data를 train data(80%)와 evaluation data(20%)로 분할하여 반환,
def train_test_split(data, ratio = 80):
    i = round(len(data)*ratio/100)
    return data[:i], data[i:]

# random하게 1 batch를 뽑아 반환
def next_random_batch(X_data, y_data, batch_size):
    idx = np.random.choice(len(y_data), batch_size, replace=False)
    return X_data[idx], y_data[idx]

# random하게 1 batch를 뽑아 반환
def next_random_interval_batch(X_data, y_data, batch_size, future_day):
    idx = np.random.choice(len(y_data) - batch_size*future_day - 1, replace=False)
    sel_idx = []
    for i in range(idx, idx + batch_size*future_day, future_day):
        sel_idx.append(i)
    return X_data[sel_idx], y_data[sel_idx]

# suffle, batch_size 설정하여 반환 : (input, target) 같이
def get_batches(data_x, data_y, SHUFFLE_BUF=1000, BATCH_SIZE=100):
    return tf.data.Dataset.from_tensor_slices((data_x, data_y)).shuffle(SHUFFLE_BUF).batch(BATCH_SIZE)

# suffle, batch_size 설정하여 반환 :  input, target 따로
def get_batches(dataset, SHUFFLE_BUF=1000, BATCH_SIZE=100):
    return tf.data.Dataset.from_tensor_slices(dataset).shuffle(SHUFFLE_BUF).batch(BATCH_SIZE)

#LSTM의 마지막 출력값 추출
def extract_last_output(output):
    last_output = np.array(output)[0:,-1].reshape(-1)
    return last_output

# 변환된 target 값을 원 상태로 복원
def back_to_price(pred_values, base_prices, conversion_type='diff'):
    if conversion_type == 'rate':
        restored_price = (pred_values/100+1)*base_prices
    elif conversion_type == 'diff':
        restored_price = list(map(operator.add, pred_values, base_prices))
    elif conversion_type == 'logdiff':
        restored_price = list(map(operator.add, np.exp(pred_values), base_prices))
    return list(restored_price)

class Evaluate(object):
    def __init__(self,prediction, true):
        self.prediction = prediction
        self.true = true
    
    def MSE(self):
        result =(sum((self.prediction-self.true)**2)/len(self.true))
        
        return round(result,4)
    def MAPE(self,predict_price, true_price):
        #true_price[np.where(true_price == 0)] = 1
        result = sum(abs((np.array(true_price)-np.array(predict_price))/np.array(true_price)))*100/len(true_price)
        return round(result,4)

    def compute_rise_fall(self):
        self.pred_rise_fall = list()
        self.output_rise_fall = list()
        for i in range(len(self.prediction)):

            if self.prediction[i] < 0 :
                self.pred_rise_fall.append(0)
            else:
                self.pred_rise_fall.append(1)

        for i in range(len(self.true)):    
            if self.true[i] < 0 :
                self.output_rise_fall.append(0)
            else:
                self.output_rise_fall.append(1)
                
        return self.pred_rise_fall, self.output_rise_fall
    
    def precision_recall(self):
        self.confus_mat = confusion_matrix(self.pred_rise_fall,self.output_rise_fall, labels=[0, 1])
        self.accu = round((self.confus_mat[0][0]+self.confus_mat[1][1])/sum(sum(self.confus_mat)),3)

        self.precision_fall = round(self.confus_mat[0][0]/sum(self.confus_mat[0]),3)
        self.recall_fall = round(self.confus_mat[0][0]/sum(self.confus_mat[:,0]),3)
        
        self.precision_rise = round(self.confus_mat[1][1]/sum(self.confus_mat[1]),3)
        self.recall_rise = round(self.confus_mat[1][1]/sum(self.confus_mat[:,1]),3)
        
        return self.confus_mat, self.accu, self.precision_fall,self.recall_fall, self.precision_rise, self.recall_rise
    
class GenerateResult():
    def __init__(self,train_predict, train_output, test_predict, test_output, test_dates, n_timestep, future_day, trans_day):
        self.train_pred = train_predict
        self.train_output = train_output
        self.test_pred = test_predict
        self.test_output = test_output

        self.n_timestep = n_timestep
        self.future_day = future_day
        self.trans_day = trans_day
        
        self.test_dates = test_dates
        
    def extract_last_output(self):
        self.train_pred = np.reshape(self.train_pred[:, -1, -1], (-1))
        self.train_output = np.reshape(self.train_output[:, -1, -1], (-1))
        self.test_pred = np.reshape(self.test_pred[:, -1, -1], (-1))
        self.test_output = np.reshape(self.test_output[:, -1, -1], (-1))

    def convert_price(self, train_base_prices, test_base_prices, conversion_type='diff'):
        # 예측값 복원
        self.train_predict_price = back_to_price(self.train_pred, train_base_prices, conversion_type )
        self.test_predict_price = back_to_price(self.test_pred, test_base_prices, conversion_type)
        # 실재값 복원
        self.train_output_price = back_to_price(self.train_output, train_base_prices, conversion_type )
        self.test_output_price = back_to_price(self.test_output, test_base_prices, conversion_type)

    def table(self):
        self.result_table = pd.DataFrame({"date":self.test_dates, "real":self.test_output.reshape(-1), "prediction": self.test_pred.reshape(-1),
                                       "real_price":self.test_output_price, "pred_price":self.test_predict_price}).reset_index(drop=True)
        
        return self.result_table

    
    def evaluation(self): #평가 지표 생성
        self.train_eval = Evaluate(self.train_pred, self.train_output)
        self.test_eval = Evaluate(self.test_pred, self.test_output)
        
        self.train_MSE = self.train_eval.MSE()
        self.test_MSE = self.test_eval.MSE()
        
        self.train_MAPE = self.train_eval.MAPE(self.train_predict_price, self.train_output_price)
        self.test_MAPE = self.test_eval.MAPE(self.test_predict_price, self.test_output_price)
        
        self.train_eval.compute_rise_fall()
        self.test_eval.compute_rise_fall()
        
        self.train_confus_mat, self.train_accu, self.train_precision_fall,self.train_recall_fall, self.train_precision_rise, self.train_recall_rise = self.train_eval.precision_recall()
        
        self.test_confus_mat, self.test_accu, self.test_precision_fall,self.test_recall_fall, self.test_precision_rise, self.test_recall_rise = self.test_eval.precision_recall()
        

    
    def save_result(self,model_name,item_name,n_units,target_type,window_size,time_interval):
        
        self.info = model_name+'_'+item_name+'_'+str(self.n_timestep)+'_'+str(time_interval)+'_'+str(self.future_day)+\
                    '_'+str(self.test_accu)
        
        
        print('info :',self.info)
        file_name = self.info+'.xlsx'
        
        self.file_path = 'result_excel/'+file_name
        self.result_table.to_excel(self.file_path)
        
        
        wb = load_workbook(self.file_path, data_only=True)
        sheet1 = wb.active
        
        sheet1.cell(1,7,'실험환경')
        sheet1.cell(2,7,'Model')
        sheet1.cell(3,7,'Item')
        sheet1.cell(4,7,'Target')
        sheet1.cell(5,7,'Timestep')
        sheet1.cell(6,7,'Interval')
        sheet1.cell(7,7,'Window size')
        sheet1.cell(8,7,'Prediction day')
        sheet1.cell(9,7,'Transfer day')
        
        
        sheet1.cell(2,8, model_name)
        sheet1.cell(3,8, item_name)
        sheet1.cell(4,8, target_type)
        sheet1.cell(5,8, self.n_timestep)
        sheet1.cell(6,8, time_interval)
        sheet1.cell(7,8, window_size)
        sheet1.cell(8,8, self.future_day)
        sheet1.cell(9,8, self.trans_day)
        
        
        sheet1.cell(1,10, '실험결과')
        sheet1.cell(2,10, 'Training Set')
        sheet1.cell(2,11, 'TRUE')
        sheet1.cell(3,10, 'PREDICTION')
        sheet1.cell(3,11, '0')
        sheet1.cell(3,12, '1')
        sheet1.cell(4,10, '0')
        sheet1.cell(5,10, '1')
        
        
        sheet1.cell(4,11,self.train_confus_mat[0][0])
        sheet1.cell(4,12,self.train_confus_mat[0][1])
        sheet1.cell(5,11,self.train_confus_mat[1][0])
        sheet1.cell(5,12,self.train_confus_mat[1][1])
        
        
        sheet1.cell(7,10, 'Test Set')
        sheet1.cell(7,11, 'TRUE')
        sheet1.cell(8,10, 'PREDICTION')
        sheet1.cell(8,11, '0')
        sheet1.cell(8,12, '1')
        sheet1.cell(9,10, '0')
        sheet1.cell(10,10, '1')
        
        
        sheet1.cell(9,11,self.test_confus_mat[0][0])
        sheet1.cell(9,12,self.test_confus_mat[0][1])
        sheet1.cell(10,11,self.test_confus_mat[1][0])
        sheet1.cell(10,12,self.test_confus_mat[1][1])
        

        
        sheet1.cell(4,13,'Rise')
        sheet1.cell(5,13,'Fall')
        sheet1.cell(3,14,'Precision')
        sheet1.cell(3,15,'Recall')
        
        sheet1.cell(4,14,self.train_precision_rise)
        sheet1.cell(4,15,self.train_recall_rise)
        
        sheet1.cell(5,14,self.train_precision_fall)
        sheet1.cell(5,15,self.train_recall_fall)
        
        
        sheet1.cell(9,13,'Rise')
        sheet1.cell(10,13,'Fall')
        sheet1.cell(8,14,'Precision')
        sheet1.cell(8,15,'Recall')
        
        sheet1.cell(9,14,self.test_precision_rise)
        sheet1.cell(9,15,self.test_recall_rise)
        
        sheet1.cell(10,14,self.test_precision_fall)
        sheet1.cell(10,15,self.test_recall_fall)
        
        
        sheet1.cell(3,18,'Training')
        sheet1.cell(3,19,'Test')
        
        sheet1.cell(4,17,'MAPE')
        sheet1.cell(5,17,'MSE')
        sheet1.cell(6,17,'Accuracy')
        
       
        print('MAPE:',self.train_MAPE)
        sheet1.cell(4,18,self.train_MAPE)
        sheet1.cell(4,19,self.test_MAPE)
        
        sheet1.cell(5,18,self.train_MSE)
        sheet1.cell(5,19,self.test_MSE)
        
        sheet1.cell(6,18,self.train_accu)
        sheet1.cell(6,19,self.test_accu)
        
        
        wb.save(self.file_path)
        return 'Succeded. Save'
    
    def save_visualization(self):
        print('MSE :', self.test_MSE,', Accuracy :',self.test_accu)
        figsize = (15,3)
        fig, ax = plt.subplots(1, 2,figsize=figsize)
        ax[0].set_title("RATIO")
        ax[0].set_xlabel("test-day")
        ax[0].set_ylabel("ratio")
        
        ax[1].set_title("PRICE")
        ax[1].set_xlabel("test-day")
        ax[1].set_ylabel("price")

        ax[0].plot(self.test_output,label="true")
        ax[0].plot(self.test_pred,label="prediction")
        ax[0].legend(loc='upper right')
        
        #ax[1].plot(self.test_output_price.reset_index(drop=True),label="true")
        #ax[1].plot(self.test_predict_price.reset_index(drop=True),label="prediction")
        ax[1].plot(self.test_output_price, label="true")
        ax[1].plot(self.test_predict_price, label="prediction")
        #ax[0].plot(self.result_ratio,label=["true","prediction"])
        #ax[1].plot(self.result_price.reset_index(drop=True))
        ax[1].legend()
        #plt.plot(self.result_df)
        #plt.plot(self.last_output)
        #return self.result_df
        #plt.show()
        filepath = 'result_pyplot/'+self.info+'.png'
        fig.savefig(filepath)
    def save_model_architecture(self,model,model_name):
        keras.utils.plot_model(model, 'model_architecture/'+model_name+'_model_with_shape_info.png', show_shapes=True)
    
    def save_model(self,model):
        model_path = 'models/'+self.info
        if not os.path.isdir(model_path): os.mkdir(model_path)
        tf.saved_model.save(model, model_path)


def predict_batch_test2(model, test_input,test_target,batch_size):
    int(model.output.shape[1])
    int(model.output.shape[2])
    pred = np.zeros((test_input.shape[0],test_input.shape[1],int(model.output.shape[2])),dtype = np.float32)
    
    #pred = list()
    for i in range(int(len(test_input)/batch_size)):
        result = model([test_input[i*batch_size:(i+1)*batch_size],test_target[i*batch_size:(i+1)*batch_size]],training = False)
        #pred += list(result)
        pred[i*batch_size:(i+1)*batch_size] = result
        print("[Succeess] batch :",(i+1)*batch_size)
    if len(test_input)%batch_size != 0:
        result = model([test_input[(i+1)*batch_size:],test_input[(i+1)*batch_size:]],training = False)
    #pred += list(result)
        pred[(i+1)*batch_size:] = result
    #pred = np.append(pred,result)
    print("Numpy Converting...")
    #pred = np.array(pred)
    print("[Success] Numpy conversion")
    return pred

def predict_batch_test(model, test_input,batch_size):
    int(model.output.shape[1])
    int(model.output.shape[2])
    pred = np.zeros((test_input.shape[0],test_input.shape[1],int(model.output.shape[2])),dtype = np.float32)
    
    #pred = list()
    for i in range(int(len(test_input)/batch_size)):
        result = model(test_input[i*batch_size:(i+1)*batch_size],training = False)
        pred[i*batch_size:(i+1)*batch_size] = result
        #print("[Succeess] batch :",(i+1)*batch_size)
    if len(test_input)%batch_size != 0:
        result = model(test_input[(i+1)*batch_size:],training = False)
        pred[(i+1)*batch_size:] = result
    #print("Numpy Converting...")
    #print("[Success] Numpy conversion")
    return pred

# 전체 batch들을 epoch수 만틈 반복
def train(model, train_dataset, test_dataset, epochs, num_batches):
    optimizer = tf.keras.optimizers.Adam(learning_rate=util.config.learning_rate)
    for epoch in range(epochs):
        for batch in range(num_batches):
            batch_x, batch_y = next(iter(train_dataset))
            gradients = models.gradient(model, 'mean_square', batch_x, batch_y)
            optimizer.apply_gradients(zip(gradients, model.trainable_varialbles))

            loss = models.loss_mean_square(model, batch_x, batch_y)
            train_MSE = models.evaluate(model, batch_x, batch_y)
            test_x, test_y = next(iter(test_dataset))
            test_MSE = models.evaluate(model, test_x, test_y)

            print('epoc :', epoch, ' loss =', loss.numpy(), ' train MSE =', train_MSE.numpy(), ' test MSE =', test_MSE.numpy())

# iteration수 만큼 batch들을 학습, 총 학습 batch 수 = iterations, 전체 데이터 반복 횟수 = iterations / (전체 데이터 개수 / batch_size)
def train(model, train_dataset, test_dataset, iterations):
    optimizer = tf.keras.optimizers.Adam(learning_rate=util.config.learning_rate)
    for i in range(iterations):
        batch_x, batch_y = next(iter(train_dataset))
        gradients = models.gradient(model, 'mean_square', batch_x, batch_y)
        optimizer.apply_gradients(zip(gradients, model.trainable_varialbles))
        if i % 1 == 0:
            loss = models.loss_mean_square(model, batch_x, batch_y)
            train_MSE = models.evaluate(model, batch_x, batch_y)
            test_x, test_y = next(iter(test_dataset))
            test_MSE = models.evaluate(model, test_x, test_y)

            print('iteration :', i, ' loss =', loss.numpy(), ' train MSE =', train_MSE.numpy(), ' test MSE =', test_MSE.numpy())

# keras의 compile, fit를 사용하여 일괄 학습
def train(model, train_x, train_y, epochs):
    model.compile(optimizer='adam',
                  loss='mean_squared_error',
                  metrics=['accuracy'])
    model.fit(train_x, train_y, epochs=epochs)

def restore_future_price(x_price, y_pred, future_day):
    restored_data = (y_pred/100+1)*x_price[:len(y_pred)]
    return restored_data
