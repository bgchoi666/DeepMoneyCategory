# Copyright 2018 Bimghi Choi. All Rights Reserved.
# ResLSTM2noise.py

# _*_ coding: utf-8 _*_

from tensorflow import keras
import tensorflow as tf

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split

#import os
#os.environ["CUDA_DEVICE_ORDER"]="PCI_BUS_ID"
#os.environ["CUDA_VISIBLE_DEVICES"]='0, 1'

import util
import preprocess as prepro
import models
import math
import datetime

from sklearn.metrics import confusion_matrix
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from numba import jit
from openpyxl import load_workbook

import os
import operator
import gc
gc.collect()

file_name = 'Data/kospi200f_36_60M_1923.csv'
item_name = 'kospi200f_36_60M_1923'

remove_columns = ['date']
target_column = '고저점'
input_columns = []
target_type = 'diff'

model_name = 'transfer_LSTM2-고저점'
channel = False

train_size = 1800
valid_size = 1800
test_size = 1923
epochs = 30
max_repeat_cnt = 5000

input_size = 36
n_unit = 100
batch_size = 20
learning_rate = 0.0005

checkpoint_path = "transfer_LSTM2-고저점/60M_input36_sample1923"
checkpoint_path_best = "transfer_LSTM2-고저점/60M_input36_sample1923_best"

# 종가를 고점 1, 저점 2, 보통 0 로 변환
dataframe = pd.read_csv(file_name, encoding='euc-kr')
df = dataframe.copy()

#strategy = tf.distribute.MirroredStrategy()
#with strategy.scope():

#model = models.LSTM_tanh(n_timestep,input_size,n_unit,regularizers_alpha=0.01,drop_rate=0.5)

model = tf.keras.Sequential([
    tf.keras.Input(shape=(input_size)),
    tf.keras.layers.Dense(n_unit, activation='relu'),
    tf.keras.layers.Dense(int(n_unit/2), activation='relu'),
    tf.keras.layers.Dense(3, activation='softmax')
])

#cp_callback = tf.keras.callbacks.ModelCheckpoint(
#    checkpoint_path, verbose=1, save_weights_only=True,
    # 다섯 번째 에포크마다 가중치를 저장합니다
#    save_freq=5)

model.compile(optimizer='adam',
              loss=keras.losses.sparse_categorical_crossentropy,
              #callbacks=[cp-callback]
              metrics=['accuracy'])
model.save_weights(checkpoint_path)

gc.collect()

# 날짜 column 제거, target column 맨 뒤로
df = df.drop(remove_columns, axis=1, inplace=False)
train_df = df.iloc[: valid_size]

def train():
    # ceate train data
    train_data = train_df.values[: train_size, :]

    # create vlaid data
    valid_data = train_df.values[train_size:, :]
    valid_x = valid_data[:, :input_size]
    valid_y = valid_data[:, input_size]

    pre_accu = 0
    repeat_cnt = 0

    while repeat_cnt < max_repeat_cnt:

        gc.collect()

        repeat_cnt += 1

        train_x = train_data[:, :input_size]
        train_y = train_data[:, input_size]

        # if train_size < valid_size:
        # train data중 50%만 사용. . .
        train_x, _, train_y, _ = train_test_split(train_x, train_y, test_size=0.5)

        # else:#if train_size >= valid_size:
        # valid data 20% 사용. . .
        train_x, valid_x, train_y, valid_y = train_test_split(train_x, train_y, test_size=0.2)

        # target 재설정하여 model training
        model.load_weights(checkpoint_path)
        early_stopping = tf.keras.callbacks.EarlyStopping(patience=2, verbose=0)
        model.fit(train_x, train_y, batch_size=batch_size, epochs=epochs, verbose=0, callbacks=[early_stopping],
                  validation_data=(valid_x, valid_y))

        prediction = model.predict(valid_x).reshape(-1, 3)
        prediction = np.argmax(prediction, axis=1).reshape(-1)

        # calculate accuracy
        c = 0
        s = len(prediction)
        for i in range(len(prediction)):
            if valid_y[i] == 0. and prediction[i] == 0.:
                c += 1
            elif valid_y[i] == 1. and prediction[i] == 1.:
                c += 1
            elif valid_y[i] == 2. and prediction[i] == 2.:
                c += 1
        accu = c / s
        if repeat_cnt % 100 == 0:
            print("반복 횟수 : " + str(repeat_cnt) + " accuracy = " + str(accu))

        if accu > pre_accu:
            best_x = train_x
            best_y = train_y
            model.fit(valid_x, valid_y, batch_size=batch_size, epochs=3, verbose=0)
            model.save_weights(checkpoint_path_best)

            print("best accuracy " + str(accu))
            pre_accu = accu

    print("best accuracy " + str(pre_accu))

    return best_x, best_y

def test(best_x, best_y):
    # create test data
    test_data = df.iloc[valid_size: test_size].values
    test_x = test_data[:, :input_size]
    test_y = test_data[:, input_size]

    # 선택 1: 최고의 데이터 셋으로 다시 학습
    # best_x = np.concatenate([best_x, valid_x])
    # best_y = np.concatenate([best_y, valid_y])

    # model.load_weights(checkpoint_path)
    # early_stopping = tf.keras.callbacks.EarlyStopping(patience=2, verbose=0)
    # model.fit(best_x, best_y, batch_size=batch_size, epochs=epochs)

    # 선택 2: 저장된 최고의 학습 weight을 reload하여 곧바로 예측
    model.load_weights(checkpoint_path_best)

    test_prediction = model.predict(test_x)
    test_prediction = np.argmax(test_prediction, axis=1).reshape(-1)

    # calculate accuracy
    c = 0
    s = len(test_prediction)
    for i in range(len(test_prediction)):
        if test_y[i] == 0 and test_prediction[i] == 0:
            c += 1
        elif test_y[i] == 1 and test_prediction[i] == 1:
            c += 1
        elif test_y[i] == 2 and test_prediction[i] == 2:
            c += 1
    print('accuracy = ', c / s)

    return test_prediction, test_y

# LSTM의 마지막 출력값 추출
def extract_last_output(output):
    last_output = np.array(output)[0:, -1].reshape(-1)
    return last_output


class Evaluate(object):
    def __init__(self, prediction, true):
        self.prediction = prediction
        self.true = true

    def MSE(self):
        result = (sum((self.prediction - self.true) ** 2) / len(self.true))

        return round(result, 4)

    def MAPE(self, predict_price, true_price):
        # true_price[np.where(true_price == 0)] = 1
        result = sum(abs((np.array(true_price) - np.array(predict_price)) / np.array(true_price))) * 100 / len(
            true_price)
        return round(result, 4)

    def compute_rise_fall(self):
        self.pred_rise_fall = list()
        self.output_rise_fall = list()
        for i in range(len(self.prediction)):

            if self.prediction[i] == 0:
                self.pred_rise_fall.append(0)
            elif self.prediction[i] == 1:
                self.pred_rise_fall.append(1)
            else:
                self.pred_rise_fall.append(2)

        for i in range(len(self.true)):
            if self.true[i] == 0:
                self.output_rise_fall.append(0)
            elif self.true[i] == 1:
                self.output_rise_fall.append(1)
            else:
                self.output_rise_fall.append(2)

        return self.pred_rise_fall, self.output_rise_fall

    def precision_recall(self):
        self.confus_mat = confusion_matrix(self.output_rise_fall, self.pred_rise_fall, labels=[0, 1, 2])
        # self.accu = round((self.confus_mat[0][0]+self.confus_mat[1][1])/sum(sum(self.confus_mat)),3)

        # self.precision_fall = round(self.confus_mat[0][0]/sum(self.confus_mat[0]),3)
        # self.recall_fall = round(self.confus_mat[0][0]/sum(self.confus_mat[:,0]),3)

        # self.precision_rise = round(self.confus_mat[1][1]/sum(self.confus_mat[1]),3)
        # self.recall_rise = round(self.confus_mat[1][1]/sum(self.confus_mat[:,1]),3)

        self.accu = accuracy_score(self.output_rise_fall, self.pred_rise_fall)

        self.precision_neutral = precision_score(self.output_rise_fall, self.pred_rise_fall, average=None)[0]
        self.precision_rise = precision_score(self.output_rise_fall, self.pred_rise_fall, average=None)[1]
        self.precision_fall = precision_score(self.output_rise_fall, self.pred_rise_fall, average=None)[2]

        self.recall_neutral = recall_score(self.output_rise_fall, self.pred_rise_fall, average=None)[0]
        self.recall_rise = recall_score(self.output_rise_fall, self.pred_rise_fall, average=None)[1]
        self.recall_fall = recall_score(self.output_rise_fall, self.pred_rise_fall, average=None)[2]

        return self.confus_mat, self.accu, self.precision_neutral, self.recall_neutral, self.precision_fall, self.recall_fall, self.precision_rise, self.recall_rise


class GenerateResult():
    def __init__(self, test_predict, test_output):
        self.test_pred = test_predict
        self.test_output = test_output

    def table(self):
        self.result_table = pd.DataFrame(
            {"real": self.test_output.reshape(-1), "prediction": self.test_pred.reshape(-1)}).reset_index(drop=True)

        return self.result_table

    def evaluation(self):  # 평가 지표 생성

        self.test_eval = Evaluate(self.test_pred, self.test_output)

        self.test_MSE = self.test_eval.MSE()

        self.test_eval.compute_rise_fall()

        self.test_confus_mat, self.test_accu, self.test_precision_neutral, self.test_recall_neutral, self.test_precision_fall, self.test_recall_fall, self.test_precision_rise, self.test_recall_rise = self.test_eval.precision_recall()

    def save_result(self, model_name, item_name):
        self.info = model_name + '_' + item_name + '_' + str(epochs) + '_' + str(self.test_accu)

        print('info :', self.info)
        file_name = self.info + '.xlsx'

        model_path =  'result_excel/' + model_name
        if not os.path.isdir(model_path): os.mkdir(model_path)
        self.file_path = model_path + "/" + file_name
        self.result_table.to_excel(self.file_path)

        wb = load_workbook(self.file_path, data_only=True)
        sheet1 = wb.active

        sheet1.cell(1, 9, '실험환경')
        sheet1.cell(2, 9, 'Model')
        sheet1.cell(3, 9, 'Item')

        sheet1.cell(2, 10, model_name)
        sheet1.cell(3, 10, item_name)

        sheet1.cell(18, 9, '실험결과')
        sheet1.cell(19, 9, 'Test Set')
        sheet1.cell(19, 10, 'TRUE')
        sheet1.cell(20, 9, 'PREDICTION')
        sheet1.cell(20, 10, '0')
        sheet1.cell(20, 11, '1')
        sheet1.cell(20, 12, '2')

        sheet1.cell(21, 9, '0')
        sheet1.cell(22, 9, '1')
        sheet1.cell(23, 9, '2')

        sheet1.cell(21, 10, self.test_confus_mat[0][0])
        sheet1.cell(21, 11, self.test_confus_mat[1][0])
        sheet1.cell(21, 12, self.test_confus_mat[2][0])

        sheet1.cell(22, 10, self.test_confus_mat[0][1])
        sheet1.cell(22, 11, self.test_confus_mat[1][1])
        sheet1.cell(22, 12, self.test_confus_mat[2][1])

        sheet1.cell(23, 10, self.test_confus_mat[0][2])
        sheet1.cell(23, 11, self.test_confus_mat[1][2])
        sheet1.cell(23, 12, self.test_confus_mat[2][2])

        sheet1.cell(25, 9, 'neutral')
        sheet1.cell(26, 9, 'rise')
        sheet1.cell(27, 9, 'Fall')
        sheet1.cell(24, 10, 'Precision')
        sheet1.cell(24, 11, 'Recall')

        sheet1.cell(25, 10, self.test_precision_neutral)
        sheet1.cell(26, 10, self.test_precision_rise)
        sheet1.cell(27, 10, self.test_precision_fall)

        sheet1.cell(25, 11, self.test_recall_neutral)
        sheet1.cell(26, 11, self.test_recall_rise)
        sheet1.cell(27, 11, self.test_recall_fall)

        sheet1.cell(29, 9, 'MSE')
        sheet1.cell(30, 9, 'Accuracy')

        sheet1.cell(29, 10, self.test_MSE)

        sheet1.cell(30, 10, self.test_accu)

        wb.save(self.file_path)
        return 'Succeded. Save'

    def save_model(self, model):
        model_path = 'models/' + self.info
        if not os.path.isdir(model_path): os.mkdir(model_path)
        tf.saved_model.save(model, model_path)

def save(test_prediction, test_y):
    result = GenerateResult(test_prediction, test_y.reshape(-1))
    result.evaluation()
    result.table()
    result.save_result(model_name, item_name)
    # result.save_visualization()
    # result.save_model(model)

def predict():
    # prediction
    pred_df0 = pd.read_csv('Data/kospi200f_36_60M_test.csv', encoding='euc-kr')
    pred_df = pred_df0.drop(remove_columns, axis=1, inplace=False)
    pred_data = pred_df.values
    pred_x = pred_data[:, :input_size]

    # model.load_weights(check_point_best)
    pred = model.predict(pred_x)
    pred = np.argmax(pred, axis=1).reshape(-1)

    print(" 1: 고점, 2: 저점")
    for i in range(len(pred)):
        print(pred_df0['date'].values[i] + ": " + str(pred[i]))

if __name__ == '__main__':
    best_x, best_y = train()
    test_prediction, test_y = test(best_x, best_y)
    save(test_prediction, test_y)